import openpyxl
import datetime
import sys

# ---------------------------------------------------------
# 便利な関数たち
# ---------------------------------------------------------
def to_wareki(year_int):
    """ 西暦 -> 令和 (簡易版) """
    reiwa_year = year_int - 2018
    if reiwa_year <= 0: return str(year_int) 
    return str(reiwa_year)

def calculate_age(born_date, today):
    """ 年齢計算 """
    try:
        return today.year - born_date.year - ((today.month, today.day) < (born_date.month, born_date.day))
    except:
        return ""

def safe_get_cell(ws, address):
    """
    【重要】最強のセル取得関数
    1. 範囲指定(A1:B2)なら左上(A1)を返す
    2. 結合セル(MergedCell)の右側や下側なら、自動的に左上の「親セル」を探して返す
    """
    try:
        # まず普通に取得してみる
        target = ws[address]

        # 1. もし「A1:B2」のような範囲(タプル)だったら、左上を取り出す
        if isinstance(target, tuple):
            target = target[0][0]

        # 2. もし「結合セルの従属セル(MergedCell)」だったら、親を探す
        # (MergedCellには値を書き込めないため)
        if "MergedCell" in str(type(target)):
            for rng in ws.merged_cells.ranges:
                # そのセルが結合範囲に含まれているか確認
                if target.coordinate in rng:
                    # 結合範囲の左上（親）のセルを返す
                    return ws.cell(row=rng.min_row, column=rng.min_col)
        
        # 普通のセルならそのまま返す
        return target

    except Exception as e:
        # print(f"セル取得エラー {address}: {e}") # デバッグ用
        return None

def mark_checkbox(ws, address):
    """ 指定されたセルの □ を ■ に変える """
    cell = safe_get_cell(ws, address)
    if cell and hasattr(cell, 'value') and isinstance(cell.value, str):
        if "□" in cell.value:
            cell.value = cell.value.replace("□", "■")

def unmark_checkbox(ws, address):
    """ 指定されたセルの ■ を □ に戻す """
    cell = safe_get_cell(ws, address)
    if cell and hasattr(cell, 'value') and isinstance(cell.value, str):
        if "■" in cell.value:
            cell.value = cell.value.replace("■", "□")

# ---------------------------------------------------------
# メイン処理
# ---------------------------------------------------------
def update_opinion_form(template_path, output_path, data):
    # data構造: { "text_data": {...}, "check_cells": [...] }
    
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        return f"エラー: テンプレート読み込み失敗: {e}"

    try:
        ws_front = wb.worksheets[0]
        ws_back = wb.worksheets[1] if len(wb.worksheets) > 1 else None
    except:
        return "エラー: シート取得失敗"

    # --- 1. 固定情報・日付・年齢の処理 ---
    today = datetime.date.today()
    
    # 自動入力データの作成
    text_updates = {
        "DH3": to_wareki(today.year),
        "DR3": str(today.month),
        "EA3": str(today.day),
        "T19": "桐生整形外科病院",
        "CN19": "0277", "CZ19": "40", "DL19": "2600",
        "T20": "桐生市相生町1丁目253-1",
        "CN20": "0277", "CZ20": "40", "DL20": "2602",
    }
    
    # 年齢計算
    birth_year = data.get("text_data", {}).get("A14")
    dob_str = data.get("meta_birth_date")
    if dob_str:
        try:
            dob_dt = datetime.datetime.strptime(dob_str, "%Y-%m-%d").date()
            age = calculate_age(dob_dt, today)
            text_updates["AT14"] = str(age)
        except: pass

    # データをマージ
    full_text_data = data.get("text_data", {})
    full_text_data.update(text_updates)

    # --- 2. テキスト書き込み ---
    # 裏面のセル番地リスト
    back_cells = ["BC8", "BX8", "A58", "X9", "Z17", "CT17", "Z19", "T23", "CR23", "AG50", "CT50", "AG51", "CT51", "AG52", "AA54"]
    
    for addr, val in full_text_data.items():
        if val:
            target_ws = ws_front
            if addr in back_cells and ws_back:
                target_ws = ws_back
            
            cell = safe_get_cell(target_ws, addr)
            if cell:
                # MergedCell対策済みのセルに書き込む
                cell.value = val

    # --- 3. チェックボックス書き込み ---
    check_list = data.get("check_cells", [])
    for addr in check_list:
        # 表にあるかもしれないし、裏にあるかもしれないので両方トライ
        mark_checkbox(ws_front, addr)
        if ws_back:
            mark_checkbox(ws_back, addr)

    try:
        wb.save(output_path)
        return "成功"
    except Exception as e:
        return f"保存エラー: {e}"

if __name__ == "__main__":
    print("main.pyのテスト実行です。Webアプリ(app.py)から実行してください。")
